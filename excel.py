import openpyxl
import json
from openpyxl.styles import numbers

# Cargar los datos de los gastos desde el archivo JSON
with open('data/gastos.json') as archivo:
    datos_gastos = json.load(archivo)

# Obtener el usuario en sesión activa
with open('data/sesion.json') as archivo_sesion:
    datos_sesion = json.load(archivo_sesion)
    email_sesion = datos_sesion['usuario']

# Filtrar los gastos por el usuario en sesión
gastos_usuario = [gasto for gasto in datos_gastos if gasto['email'] == email_sesion]

# Crear un nuevo archivo de Excel
wb = openpyxl.Workbook()

# Obtener la hoja activa
hoja = wb.active

# Escribir los encabezados de la tabla
hoja["A1"] = "Usuario"
hoja["B1"] = "Fecha"
hoja["C1"] = "Categoría"
hoja["D1"] = "Monto"

# Dar formato de moneda a la columna de Monto
currency_format = numbers.FORMAT_CURRENCY_USD
for fila, gasto in enumerate(gastos_usuario, start=2):
    hoja.cell(row=fila, column=1).value = gasto["fecha"]
    hoja.cell(row=fila, column=2).value = gasto["fecha"]
    hoja.cell(row=fila, column=3).value = gasto["categoria"]
    monto_celda = hoja.cell(row=fila, column=4)
    monto_celda.value = float(gasto["monto"])
    monto_celda.number_format = currency_format

# Ajustar el ancho de las columnas
hoja.column_dimensions["A"].width = 12
hoja.column_dimensions["B"].width = 12
hoja.column_dimensions["C"].width = 12
# Guardar el archivo de Excel
wb.save("data/gastos_usuario.xlsx")
